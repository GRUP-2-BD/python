import zipfile
import os
import pandas as pd # pip install pandas / pip install xlrd
import openpyxl as xl
import mysql.connector
import datetime

dirActual = os.path.dirname(__file__)
print (dirActual)
nomFitxerZip = "02_197706_1.zip"
dirUnzip = os.path.join(os.path.dirname(__file__) , "excels")
pathFitxerZip = os.path.join(dirActual,nomFitxerZip)
print (dirUnzip)

# Extraiem el contingut de pathFitxerZip a dirUnzip
with zipfile.ZipFile(pathFitxerZip, 'r') as zipRef:
 zipRef.extractall(dirUnzip)


#Tractar un fitxer Excel
dirActual = os.path.dirname(__file__)+"\excels"
nomFitxerXls = "02_197706_1.xlsx"
pathFitxerXls = os.path.join(dirActual,nomFitxerXls)
print(pathFitxerXls)

wb = xl.load_workbook(pathFitxerXls, read_only=True)
sheet = wb["Municipios"]
#print (sheet.max_row) #retorna la quantiat màxima de files
#print (sheet.max_column) #retorna la quantitat màixma de columnes
"""
fila 7 comencen les dades
Columnes:
   1: Nom de la comunitat
   2: Codi de Provincia
   3: Nom de la Provincia
   4: Codi de Municipi
   5: Nom de Municipi
   6: Població
   7: Número de meses
   8: Total del cens electoral
   9: Total de vots
   10: Vots vàlids
   11: Vots a candidatures
   12: Vots en blanc
   13: Vots nuls
   14: shee.max_column (partits polítics)

"""


cnx = mysql.connector.connect(host='192.168.56.105',user='perepi',password='pastanaga', database='mydb')
cursor = cnx.cursor()

avui = datetime.datetime.now().date()
stm_insert_empleat = ("INSERT INTO comunitats_autonomes "
                "(nom,codi_ine) "
                "VALUES (%s, %s)")

comunitats=[]

for i in range(7,8121):
    if sheet.cell(row=i,column=1).value not in comunitats:
        dades_empleat = ((sheet.cell(row=i,column=1).value).rstrip(" "),sheet.cell(row=i,column=2).value)
        cursor.execute(stm_insert_empleat, dades_empleat)
        comunitats.append(sheet.cell(row=i,column=1).value)
        print("Inserit! {0}".format(dades_empleat))
    else:
        print(i)

        #print(comunitats)
# Executem l'INSERT
# Si la taula tenia un valor autoincremental aquest es pot recollir mitjançant lastrowid o _last_insert_id.
empleat_id = cursor.lastrowid
print(cursor._last_insert_id)
cnx.commit()
cursor.close()
cnx.close()

